"""
Script principal – Trabajo 3: Búsqueda en Vecindarios para NWJSSP
===================================================================
Ejecuta dos modos según la variable MODE al inicio del archivo:

    MODE = "final"      → Corre la configuración definitiva de cada algoritmo
                          y guarda los Excel de entrega.

    MODE = "parametric" → Corre análisis comparativo de parámetros para VND
                          y para MS-LNS-SA, guardando un Excel por configuración.

Algoritmos:
    1. VND       – Variable Neighborhood Descent (N1=2-opt, N2=Swap-10, N3=Insertion-10)
    2. MS-LNS-SA – Multi-Start + LNS + Recocido Simulado
"""

import os
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from constructive import ConstructiveAlgorithm
from vnd import VNDSearch
from ms_lns_sa import MSLNSSearch
from read_instances import read_nwjssp_instance

# ===========================================================================
# ════════════════════════  CONFIGURACIÓN PRINCIPAL  ════════════════════════
# ===========================================================================

# Modo de ejecución: "final" | "parametric"
MODE = "final"
# En modo final puede ejecutarse solo un algoritmo si se especifica:
#   FINAL_ALGO_MODE=vnd      -> solo VND
#   FINAL_ALGO_MODE=mslns    -> solo MS-LNS-SA
#   FINAL_ALGO_MODE=both     -> ambos (comportamiento por defecto)
FINAL_ALGO_MODE = os.getenv("FINAL_ALGO_MODE", "mslns").lower()

# Directorio con los archivos .txt de instancias
INSTANCES_DIR = os.getenv("INSTANCES_DIR", "instances")
INSTANCES_DIR_PARA = os.getenv("INSTANCES_DIR_PARA", "instances_parametric")

# Tiempo límite POR EJECUCIÓN de algoritmo (segundos)
TIME_LIMIT = 3600   # 1 hora  (modo final)
TIME_LIMIT_PARAM = 300  # 5 min por configuración en análisis paramétrico
                        # (ajustar según cuántas instancias/configs se quieran probar)

# ---------------------------------------------------------------------------
# Parámetros definitivos (usados en MODE="final")
# ---------------------------------------------------------------------------
VND_PARAMS_FINAL = dict(max_range=10, improve_n1="BI", improve_n2="BI", improve_n3="FI")

MSLNS_PARAMS_FINAL = dict(
    max_range=10,
    n_starts=5,
    q_init=4,
    q_max=10,
    no_improve_q_step=15,
    sa_t0=None,
    sa_cooling=0.995,
    seed=42,
)

# ---------------------------------------------------------------------------
# Configuraciones paramétricas a comparar (usadas en MODE="parametric")
#
# VND: el único parámetro relevante es max_range (radio del vecindario Swap e Insertion).
# MS-LNS-SA: se analizan n_starts, q_init y sa_cooling.
# ---------------------------------------------------------------------------

VND_PARAM_CONFIGS = [
    # ── Variación de max_range (estrategias fijas: BI, BI, FI) ──────────────
    # Permite ver el efecto del radio del vecindario en la calidad de la solución
    ("VND_r05_BI_BI_FI", dict(max_range=5,  improve_n1="BI", improve_n2="BI", improve_n3="FI")),
    ("VND_r10_BI_BI_FI", dict(max_range=10, improve_n1="BI", improve_n2="BI", improve_n3="FI")),  # <-- final
    ("VND_r20_BI_BI_FI", dict(max_range=20, improve_n1="BI", improve_n2="BI", improve_n3="FI")),

    # ── Variación de la estrategia de mejora (max_range=10 fijo) ────────────
    # Permite comparar BI vs FI en cada vecindario individualmente

    # N1 (2-opt): FI vs BI  (N2 y N3 fijos)
    ("VND_r10_FI_BI_FI", dict(max_range=10, improve_n1="FI", improve_n2="BI", improve_n3="FI")),

    # N2 (Swap): FI vs BI  (N1 y N3 fijos)
    ("VND_r10_BI_FI_FI", dict(max_range=10, improve_n1="BI", improve_n2="FI", improve_n3="FI")),

    # N3 (Insertion): BI vs FI  (N1 y N2 fijos)
    ("VND_r10_BI_BI_BI", dict(max_range=10, improve_n1="BI", improve_n2="BI", improve_n3="BI")),

    # Todas FI (exploración rápida) vs todas BI (exploración exhaustiva)
    ("VND_r10_FI_FI_FI", dict(max_range=10, improve_n1="FI", improve_n2="FI", improve_n3="FI")),
    ("VND_r10_BI_BI_BI", dict(max_range=10, improve_n1="BI", improve_n2="BI", improve_n3="BI")),
]

MSLNS_PARAM_CONFIGS = [
    # ── Variación de n_starts (cuántas soluciones iniciales) ─────────────────
    ("MSLNS_starts3_q4_c0995",  dict(n_starts=3,  q_init=4, q_max=10, no_improve_q_step=15, sa_cooling=0.995, seed=42)),
    ("MSLNS_starts5_q4_c0995",  dict(n_starts=5,  q_init=4, q_max=10, no_improve_q_step=15, sa_cooling=0.995, seed=42)),  # <-- final
    ("MSLNS_starts10_q4_c0995", dict(n_starts=10, q_init=4, q_max=10, no_improve_q_step=15, sa_cooling=0.995, seed=42)),

    # ── Variación de q_init (tamaño inicial de destrucción) ──────────────────
    ("MSLNS_starts5_q2_c0995",  dict(n_starts=5, q_init=2,  q_max=8,  no_improve_q_step=15, sa_cooling=0.995, seed=42)),
    ("MSLNS_starts5_q6_c0995",  dict(n_starts=5, q_init=6,  q_max=12, no_improve_q_step=15, sa_cooling=0.995, seed=42)),

    # ── Variación del enfriamiento SA ─────────────────────────────────────────
    ("MSLNS_starts5_q4_c0990",  dict(n_starts=5, q_init=4,  q_max=10, no_improve_q_step=15, sa_cooling=0.990, seed=42)),
    ("MSLNS_starts5_q4_c0999",  dict(n_starts=5, q_init=4,  q_max=10, no_improve_q_step=15, sa_cooling=0.999, seed=42)),
]

# ===========================================================================


# ---------------------------------------------------------------------------
# Utilidades Excel
# ---------------------------------------------------------------------------

def create_results_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def get_column_letter(col_num):
    col_letter = ""
    col = col_num
    while col >= 0:
        col_letter = chr(65 + (col % 26)) + col_letter
        col = col // 26 - 1
        if col < 0:
            break
    return col_letter


def add_results_sheet(workbook, instance_name, flow_time, computation_time, job_start_times):
    ws = workbook.create_sheet(instance_name)
    ws["A1"] = int(flow_time)
    ws["B1"] = int(round(computation_time))

    for idx, start_time in enumerate(job_start_times):
        ws[f"{get_column_letter(idx)}2"] = int(start_time)

    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for cell in [ws["A1"], ws["B1"]]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    for idx in range(len(job_start_times)):
        ws.column_dimensions[get_column_letter(idx)].width = 12


# ---------------------------------------------------------------------------
# Helpers de ejecución
# ---------------------------------------------------------------------------

def _initial_solution(n, m, operations, release_dates):
    c = ConstructiveAlgorithm(n, m, operations, release_dates)
    sol, _, _ = c.solve()
    return sol


def run_vnd(n, m, operations, release_dates, time_limit=TIME_LIMIT, **params):
    sol0 = _initial_solution(n, m, operations, release_dates)
    algo = VNDSearch(n, m, operations, release_dates, time_limit=time_limit, **params)
    starts, flow, comp, n_solutions = algo.solve(initial_solution=sol0)
    return starts, flow, comp, n_solutions


def run_mslns(n, m, operations, release_dates, time_limit=TIME_LIMIT, **params):
    sol0 = _initial_solution(n, m, operations, release_dates)
    algo = MSLNSSearch(n, m, operations, release_dates, time_limit=time_limit, **params)
    return algo.solve(initial_solution=sol0)   # devuelve (starts, flow, comp, n_solutions)


def process_instance(instance_file, algo_fn, time_limit, params):
    try:
        n, m, ops, rd, _ = read_nwjssp_instance(instance_file)
        starts, ft, ct, n_solutions = algo_fn(n, m, ops, rd, time_limit=time_limit, **params)
        name = os.path.splitext(os.path.basename(instance_file))[0]
        return name, starts, ft, ct, n_solutions
    except Exception as e:
        print(f"  Error procesando {instance_file}: {e}")
        return None


def run_batch(instance_files, algo_fn, label, output_file, time_limit, params):
    """Ejecuta algo_fn sobre todas las instancias y guarda Excel."""
    print(f"\n  [{label}]")
    print("  " + "-" * 70)

    wb = create_results_workbook()
    total_time = 0
    total_calls = 0
    count = 0

    for f in instance_files:
        result = process_instance(f, algo_fn, time_limit, params)
        if result:
            name, sol, ft, ct, n_solutions = result
            sol_str = f"{n_solutions:8d}" if n_solutions is not None else "     N/A"
            print(f"    {name:28s} | Z={ft:12.0f} | t={ct:10.2f}ms | soluciones={sol_str}")
            total_time += ct
            total_calls += n_solutions if n_solutions is not None else 0
            count += 1
            add_results_sheet(wb, name, ft, ct, sol)

    wb.save(output_file)
    avg = total_time / count if count > 0 else 0
    avg_sol = total_calls / count if count > 0 else 0
    print(f"\n  ✓ {output_file}  (inst={count}, total={total_time/1000:.2f}s, avg={avg:.2f}ms, avg_soluciones={avg_sol:.1f})")
    return total_time, count


# ---------------------------------------------------------------------------
# Modo FINAL
# ---------------------------------------------------------------------------

def run_final(instance_files):
    print("\n" + "=" * 75)
    print("MODO FINAL  –  Configuraciones definitivas")
    print(f"Algoritmo(s): {FINAL_ALGO_MODE.upper()}")
    print("=" * 75)

    configs = []
    if FINAL_ALGO_MODE in ("both", "all", ""):
        configs = [
            (run_vnd,      "VND  (N1=2opt-BI, N2=Swap10-BI, N3=Insertion10-FI)",
             "NWJSSP_ArturoMurgueytio_VND.xlsx",      TIME_LIMIT, VND_PARAMS_FINAL),
            (run_mslns,    "MS-LNS-SA (Multi-Start + LNS + Recocido Simulado)",
             "NWJSSP_ArturoMurgueytio_MSLNS.xlsx",    TIME_LIMIT, MSLNS_PARAMS_FINAL),
        ]
    elif FINAL_ALGO_MODE == "vnd":
        configs = [
            (run_vnd,      "VND  (N1=2opt-BI, N2=Swap10-BI, N3=Insertion10-FI)",
             "NWJSSP_ArturoMurgueytio_VND.xlsx",      TIME_LIMIT, VND_PARAMS_FINAL),
        ]
    elif FINAL_ALGO_MODE in ("mslns", "ms_lns"):
        configs = [
            (run_mslns,    "MS-LNS-SA (Multi-Start + LNS + Recocido Simulado)",
             "NWJSSP_ArturoMurgueytio_MSLNS.xlsx",    TIME_LIMIT, MSLNS_PARAMS_FINAL),
        ]
    else:
        print(f"Advertencia: FINAL_ALGO_MODE='{FINAL_ALGO_MODE}' no reconocido. Ejecutando ambos algoritmos.")
        configs = [
            (run_vnd,      "VND  (N1=2opt-BI, N2=Swap10-BI, N3=Insertion10-FI)",
             "NWJSSP_ArturoMurgueytio_VND.xlsx",      TIME_LIMIT, VND_PARAMS_FINAL),
            (run_mslns,    "MS-LNS-SA (Multi-Start + LNS + Recocido Simulado)",
             "NWJSSP_ArturoMurgueytio_MSLNS.xlsx",    TIME_LIMIT, MSLNS_PARAMS_FINAL),
        ]

    summary = []
    for fn, label, outfile, tl, params in configs:
        total, count = run_batch(instance_files, fn, label, outfile, tl, params)
        summary.append((label, count, total))

    _print_summary(summary)
    return summary


# ---------------------------------------------------------------------------
# Modo PARAMÉTRICO
# ---------------------------------------------------------------------------

def run_parametric(instance_files):
    print("\n" + "=" * 75)
    print("MODO PARAMÉTRICO  –  Análisis comparativo de parámetros")
    print(f"Tiempo límite por configuración: {TIME_LIMIT_PARAM} s")
    print("=" * 75)

    all_results = []

    # ── VND ──────────────────────────────────────────────────────────────
    print("\n▸ VND  –  variando max_range y estrategia BI/FI por vecindario")
    for label, params in VND_PARAM_CONFIGS:
        outfile = f"NWJSSP_ArturoMurgueytio_{label}.xlsx"
        total, count = run_batch(
            instance_files, run_vnd, label, outfile, TIME_LIMIT_PARAM, params
        )
        all_results.append((label, count, total))

    # ── MS-LNS-SA ────────────────────────────────────────────────────────
    print("\n▸ MS-LNS-SA  –  variando n_starts, q_init, sa_cooling")
    for label, params in MSLNS_PARAM_CONFIGS:
        outfile = f"NWJSSP_ArturoMurgueytio_{label}.xlsx"
        total, count = run_batch(
            instance_files, run_mslns, label, outfile, TIME_LIMIT_PARAM, params
        )
        all_results.append((label, count, total))

    _print_summary(all_results)
    return all_results


# ---------------------------------------------------------------------------
# Resumen impreso
# ---------------------------------------------------------------------------

def _print_summary(results):
    print("\n" + "=" * 75)
    print("RESUMEN")
    print("=" * 75)
    print(f"  {'Configuración':<40} | {'Inst.':>5} | {'Total':>9} | {'Promedio':>10}")
    print("  " + "-" * 70)
    for label, count, total in results:
        avg = total / count if count > 0 else 0
        print(f"  {label:<40} | {count:>5} | {total/1000:>6.2f}s  | {avg:>8.2f}ms")
    print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    
    if MODE == "final":
        print("=" * 75)
        print("NWJSSP – Trabajo 3: VND  |  MS-LNS-SA (Multi-Start + LNS + SA)")
        print(f"Modo: {MODE.upper()}")
        print("=" * 75)

        instance_files = sorted(glob.glob(os.path.join(INSTANCES_DIR, "*.txt")))
        if not instance_files:
            print(f"Error: No se encontraron instancias en '{INSTANCES_DIR}'")
            return

        print(f"\nInstancias encontradas: {len(instance_files)}")

        run_final(instance_files)
    elif MODE == "parametric":
        instance_files = sorted(glob.glob(os.path.join(INSTANCES_DIR_PARA, "*.txt")))
        if not instance_files:
            print(f"Error: No se encontraron instancias en '{INSTANCES_DIR_PARA}'")
            return
        run_parametric(instance_files)
    else:
        print(f"Error: MODE='{MODE}' no reconocido. Use 'final' o 'parametric'.")

    print("✓ Ejecución completada\n")


if __name__ == "__main__":
    main()